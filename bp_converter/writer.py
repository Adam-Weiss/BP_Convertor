from __future__ import annotations

import csv
from datetime import datetime
from typing import Iterable, List

from .models import ConversionStats, Measurement

SMARTBP_HEADER = [
    "Date",
    "Systolic(mmHg)",
    "Diastolic(mmHg)",
    "Pulse(BPM)",
    "Weight(kgs)",
    "Pulse pressure(mmHg)",
    "MAP( mmHg)",
    "Notes",
    "Tags",
    "csRecordId",
    "source",
]

ALLOWED_COLUMNS = [
    "datetime", "sys", "dia", "pulse", "weight", "pp", "map", "notes", "tags", "source",
    "original_date", "original_time", "warnings", "corrections_applied",
]


def fmt_smartbp_datetime(dt: datetime) -> str:
    return dt.strftime("%m/%d/%y %H:%M:%S")


def _str_or_blank(value: object) -> str:
    return "" if value is None else str(value)


def write_smartbp_csv(path: str, measurements: List[Measurement]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(
            f,
            delimiter=",",
            quotechar='"',
            quoting=csv.QUOTE_ALL,
            lineterminator="\r\n",
            doublequote=True,
        )
        writer.writerow(SMARTBP_HEADER)
        for m in measurements:
            weight_val = "NA" if m.weight is None else (str(int(m.weight)) if float(m.weight).is_integer() else str(m.weight))
            row = [
                fmt_smartbp_datetime(m.datetime),
                str(m.sys),
                str(m.dia),
                _str_or_blank(m.pulse),
                weight_val,
                _str_or_blank(m.pp),
                _str_or_blank(m.map),
                m.notes.replace("\n", "\r\n"),
                m.tags,
                "",
                str(m.source),
            ]
            writer.writerow(row)


def _measurement_to_dict(m: Measurement) -> dict:
    return {
        "datetime": m.datetime,
        "sys": m.sys,
        "dia": m.dia,
        "pulse": m.pulse,
        "weight": m.weight,
        "pp": m.pp,
        "map": m.map,
        "notes": m.notes,
        "tags": m.tags,
        "source": m.source,
        "original_date": m.original_date,
        "original_time": m.original_time,
        "warnings": " | ".join(m.warnings),
        "corrections_applied": " | ".join(m.corrections_applied),
    }


def write_normalized_xlsx(path: str, measurements: List[Measurement], columns: List[str], stats: ConversionStats | None = None) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Measurements"

    ws.append(columns)
    for m in measurements:
        rowd = _measurement_to_dict(m)
        ws.append([rowd[col] for col in columns])

    if "datetime" in columns:
        cidx = columns.index("datetime") + 1
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=cidx).number_format = "yyyy-mm-dd hh:mm:ss"

    if stats is not None:
        st = wb.create_sheet("Stats")
        st.append(["metric", "value"])
        for field_name, value in stats.__dict__.items():
            st.append([field_name, value])
            if isinstance(value, datetime):
                st.cell(row=st.max_row, column=2).number_format = "yyyy-mm-dd hh:mm:ss"

    wb.save(path)
