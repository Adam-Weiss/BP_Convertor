from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from .role_inference import infer_roles


@dataclass
class DetectedTable:
    headers: List[Any]
    rows: List[List[Any]]
    roles: Dict[str, int]
    score: int


def _is_number(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    if text == "":
        return False
    try:
        float(text)
        return True
    except ValueError:
        return False


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _score_candidate(headers: List[Any], rows: List[List[Any]], roles: Dict[str, int]) -> int:
    """Score a candidate BP table block.

    Higher scores indicate stronger confidence that a row block is an actual BP
    measurements table instead of descriptive report text.
    """
    score = 0
    if {"sys", "dia"}.issubset(roles):
        score += 40
    if "datetime" in roles or "date" in roles:
        score += 25
    if "time" in roles:
        score += 8
    if "pulse" in roles:
        score += 6

    if len(rows) >= 3:
        score += min(len(rows), 40)

    sample = rows[: min(25, len(rows))]
    valid_bp_rows = 0
    monotonic_rows = 0
    pulse_like_rows = 0
    for row in sample:
        if "sys" in roles and roles["sys"] < len(row) and "dia" in roles and roles["dia"] < len(row):
            sys_v = _to_float(row[roles["sys"]])
            dia_v = _to_float(row[roles["dia"]])
            if sys_v is not None and dia_v is not None:
                if 40 <= dia_v <= 180 and 60 <= sys_v <= 280:
                    valid_bp_rows += 1
                if sys_v > dia_v:
                    monotonic_rows += 1
        if "pulse" in roles and roles["pulse"] < len(row):
            pulse_v = _to_float(row[roles["pulse"]])
            if pulse_v is not None and 25 <= pulse_v <= 260:
                pulse_like_rows += 1

    score += valid_bp_rows * 4
    score += monotonic_rows * 2
    score += pulse_like_rows

    return score


def _extract_contiguous_rows(source_rows: List[List[Any]], start_idx: int, date_col: Optional[int]) -> List[List[Any]]:
    out: List[List[Any]] = []
    empty_tolerance = 0
    for row in source_rows[start_idx:]:
        is_blank_row = all((str(cell).strip() == "" for cell in row))
        if is_blank_row:
            empty_tolerance += 1
            if empty_tolerance >= 2:
                break
            continue
        empty_tolerance = 0

        if date_col is not None and date_col < len(row):
            if str(row[date_col]).strip() == "" and len(out) >= 2:
                break
        out.append(row)
    return out


def detect_xlsx_table(input_path: str) -> DetectedTable:
    """Detect the most probable BP table in an XLSX workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(input_path, data_only=True)
    best: Optional[DetectedTable] = None
    for sheet in wb.worksheets:
        all_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        scan_rows = min(len(all_rows), 180)
        for idx in range(scan_rows):
            headers = all_rows[idx]
            roles = infer_roles(headers)
            if not ({"sys", "dia"}.issubset(roles) and ("date" in roles or "datetime" in roles)):
                continue
            date_col = roles.get("datetime", roles.get("date"))
            block = _extract_contiguous_rows(all_rows, idx + 1, date_col)
            if len(block) < 1:
                continue
            score = _score_candidate(headers, block, roles)
            cand = DetectedTable(headers=headers, rows=block, roles=roles, score=score)
            if best is None or cand.score > best.score:
                best = cand
    if best is None:
        raise ValueError(
            "Could not detect a blood-pressure table in XLSX input. "
            "Please verify the file has columns for date/datetime, systolic, and diastolic values."
        )
    return best


def _sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";"])
        return dialect.delimiter
    except csv.Error:
        if "\t" in sample:
            return "\t"
        if ";" in sample:
            return ";"
        return ","


def _read_delimited_rows(input_path: str) -> Tuple[List[List[str]], str]:
    with open(input_path, "r", encoding="utf-8", errors="replace", newline="") as handle:
        sample = handle.read(4000)
        delimiter = _sniff_delimiter(sample)
        handle.seek(0)
        reader = csv.reader(handle, delimiter=delimiter)
        rows = list(reader)
    return rows, delimiter


def detect_delimited_table(input_path: str) -> Tuple[DetectedTable, str]:
    """Detect the most probable BP table in CSV/TSV input."""
    rows, delimiter = _read_delimited_rows(input_path)

    best: Optional[DetectedTable] = None
    scan_rows = min(len(rows), 260)
    for idx in range(scan_rows):
        headers = rows[idx]
        roles = infer_roles(headers)
        if not ({"sys", "dia"}.issubset(roles) and ("date" in roles or "datetime" in roles)):
            continue
        date_col = roles.get("datetime", roles.get("date"))
        block = _extract_contiguous_rows(rows, idx + 1, date_col)
        if len(block) < 1:
            continue
        score = _score_candidate(headers, block, roles)
        cand = DetectedTable(headers=headers, rows=block, roles=roles, score=score)
        if best is None or cand.score > best.score:
            best = cand

    if best is None:
        raise ValueError(
            "Could not detect a blood-pressure table in delimited text. "
            "Ensure the file contains a header row with date/datetime plus systolic and diastolic columns."
        )
    return best, delimiter


def detect_table(input_path: str) -> DetectedTable:
    """Detect and return the BP-like table from a supported input file."""
    suffix = Path(input_path).suffix.lower()
    if suffix == ".xlsx":
        return detect_xlsx_table(input_path)
    if suffix in {".csv", ".tsv"}:
        table, _ = detect_delimited_table(input_path)
        return table
    raise ValueError(f"Unsupported input extension: {suffix}. Supported: .xlsx, .csv, .tsv")
