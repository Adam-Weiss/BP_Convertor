import argparse
import csv
import re
from datetime import datetime, date, time
from typing import Any, Optional, Tuple, List, Dict

import pandas as pd
from openpyxl import load_workbook


SMARTBP_HEADER = [
    "Date",
    "Systolic(mmHg)",
    "Diastolic(mmHg)",
    "Pulse(BPM)",
    "Weight(kgs)",
    "Pulse pressure(mmHg)",
    "MAP( mmHg)",
    "Notes",
    "Tags",
    "csRecordId",
    "source",
]

# Canonical "input" header concepts we accept for detection (synonyms)
HEADER_SYNONYMS: Dict[str, List[str]] = {
    "date": ["date"],
    "sys": ["sys (mmhg)", "sys", "systolic(mmhg)", "systolic", "systolic mmhg"],
    "dia": ["dia (mmhg)", "dia", "diastolic(mmhg)", "diastolic", "diastolic mmhg"],
    "pulse": ["pulse (bpm)", "pulse(bpm)", "pulse", "hr", "heart rate"],
    "weight": ["weight kgs", "weight (kgs)", "weight(kgs)", "weight kgs ", "weight(kg)", "weight kg", "weight(kgs)", "weight"],
    "pp": ["pp (mmhg)", "pp", "pulse pressure(mmhg)", "pulse pressure", "pp\n(mmhg)"],
    "map": ["map (mmhg)", "map", "map( mmhg)", "map\n(mmhg)"],
    "notes": ["notes", "note"],
    "tags": ["tags", "tag"],
}

# For your specific sheet style (with line breaks in headers)
# Normalize newlines to spaces during matching, so it will still match.
def _norm_header(s: Any) -> str:
    if s is None:
        return ""
    x = str(s)
    x = x.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    x = re.sub(r"\s+", " ", x).strip().lower()
    return x

def _norm_cell_text(x: Any) -> str:
    if x is None:
        return ""
    s = str(x)
    s = s.replace("\r\n", " ").replace("\n", "").replace("\r", "")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _is_numberish(x: Any) -> bool:
    if x is None:
        return False
    s = _norm_cell_text(x)
    if s == "" or s.upper() == "NA":
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False

def _parse_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date) and not isinstance(value, datetime):
        dt = datetime(value.year, value.month, value.day, 0, 0, 0)
    else:
        s = _norm_cell_text(value)
        if not s:
            raise ValueError("Empty Date value")

        fmts = [
            "%m/%d/%y %H:%M:%S",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%y %H:%M",
            "%m/%d/%Y %H:%M",
            "%d-%b-%y %H:%M:%S",
            "%d-%b-%y %H:%M",
            "%d-%b-%Y %H:%M:%S",
            "%d-%b-%Y %H:%M",
            "%d-%b-%y %H:%M:%S",
            "%d-%b-%y %H:%M",
        ]
        dt = None
        for f in fmts:
            try:
                dt = datetime.strptime(s, f)
                break
            except ValueError:
                continue
        if dt is None:
            # heuristic: M/D/YY H:MM(:SS)?
            m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4}) (\d{1,2}):(\d{2})(?::(\d{2}))?$", s)
            if m:
                mm, dd, yy, hh, mi, ss = m.groups()
                year = int(yy)
                if year < 100:
                    year += 2000
                dt = datetime(year, int(mm), int(dd), int(hh), int(mi), int(ss or 0))
            else:
                raise ValueError(f"Unrecognized Date format: {s!r}")

    return dt.replace(second=int(dt.second or 0))

def _fmt_smartbp_datetime(dt: datetime) -> str:
    return dt.strftime("%m/%d/%y %H:%M:%S")

def _to_int_or_blank(x: Any) -> str:
    s = _norm_cell_text(x)
    if s == "" or s.lower() == "nan":
        return ""
    try:
        return str(int(float(s)))
    except ValueError:
        return s

def _to_float_or_blank(x: Any) -> str:
    s = _norm_cell_text(x)
    if s == "" or s.lower() == "nan":
        return ""
    if s.upper() == "NA":
        return "NA"
    try:
        v = float(s)
        if v.is_integer():
            return str(int(v))
        return str(v)
    except ValueError:
        return s

def _split_tags(tag_cell: Any) -> str:
    s = _norm_cell_text(tag_cell)
    if s in ("", "0"):
        return ""
    parts = [p.strip() for p in s.split(",") if p.strip()]
    return ",".join(parts)

def _build_notes(tags: str, user_notes: str, extra_note: str) -> str:
    user_notes = _norm_cell_text(user_notes)
    extra_note = _norm_cell_text(extra_note)

    # Treat placeholders as empty
    if user_notes == "0":
        user_notes = ""

    combined = user_notes
    if extra_note:
        combined = (combined + " | " if combined else "") + extra_note

    if tags:
        if combined:
            return f"Tags: {tags}\r\n\r\nNotes: {combined}"
        return f"Tags: {tags}"
    else:
        if combined:
            return f"Notes: {combined}"
        return ""

def _check_and_fix_sys_dia(sys_v: str, dia_v: str, pp_tbl: str) -> Tuple[str, str, str, str]:
    extra_notes = []

    def _as_num(s: str) -> Optional[float]:
        if not s or s.upper() == "NA":
            return None
        try:
            return float(s)
        except ValueError:
            return None

    sys_n = _as_num(sys_v)
    dia_n = _as_num(dia_v)
    pp_tbl_n = _as_num(pp_tbl)

    pp_calc_n = None
    if sys_n is not None and dia_n is not None:
        pp_calc_n = sys_n - dia_n

    if pp_calc_n is not None and pp_calc_n < 0:
        old_sys, old_dia = sys_v, dia_v
        sys_v, dia_v = dia_v, sys_v
        sys_n, dia_n = dia_n, sys_n
        pp_calc_n = sys_n - dia_n if (sys_n is not None and dia_n is not None) else None
        extra_notes.append(f"corrected swapped Sys/Dia (was Sys={old_sys}, Dia={old_dia})")

    if pp_tbl_n is not None and pp_tbl_n < 0:
        extra_notes.append(f"input PP was negative ({pp_tbl})")

    if pp_tbl_n is not None and pp_calc_n is not None:
        if abs(pp_tbl_n - pp_calc_n) >= 2:
            extra_notes.append(f"input PP={pp_tbl} differs from calc PP={int(round(pp_calc_n))}")

    pp_out = ""
    if pp_calc_n is not None:
        pp_out = str(int(round(pp_calc_n)))

    return sys_v, dia_v, pp_out, " | ".join(extra_notes)

def _match_header_concept(cell: Any, concept: str) -> bool:
    h = _norm_header(cell)
    if not h:
        return False
    for syn in HEADER_SYNONYMS[concept]:
        if h == syn:
            return True
    return False

def find_smartbp_table(ws, max_scan_rows: int = 80) -> Tuple[int, Dict[str, int]]:
    """
    Find a header row that looks like a SmartBP table.
    Returns (header_row_index_1based, column_index_map_1based)
    Raises if not found.
    """
    max_row = min(ws.max_row, max_scan_rows)
    max_col = min(ws.max_column, 60)

    for r in range(1, max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        # Build header map for this candidate row
        col_map: Dict[str, int] = {}

        for c, v in enumerate(row_vals, start=1):
            if _match_header_concept(v, "date"):
                col_map["date"] = c
            elif _match_header_concept(v, "sys"):
                col_map["sys"] = c
            elif _match_header_concept(v, "dia"):
                col_map["dia"] = c
            elif _match_header_concept(v, "pulse"):
                col_map["pulse"] = c
            elif _match_header_concept(v, "weight"):
                col_map["weight"] = c
            elif _match_header_concept(v, "pp"):
                col_map["pp"] = c
            elif _match_header_concept(v, "map"):
                col_map["map"] = c
            elif _match_header_concept(v, "notes"):
                col_map["notes"] = c
            elif _match_header_concept(v, "tags"):
                col_map["tags"] = c

        # Minimal required headers to call it "SmartBP-like"
        required = {"date", "sys", "dia", "pulse"}
        if not required.issubset(col_map.keys()):
            continue

        # Sanity check: next 1–3 rows should have numeric sys/dia/pulse
        good_rows = 0
        for rr in range(r + 1, min(r + 4, ws.max_row + 1)):
            sys_v = ws.cell(rr, col_map["sys"]).value
            dia_v = ws.cell(rr, col_map["dia"]).value
            pul_v = ws.cell(rr, col_map["pulse"]).value
            if _is_numberish(sys_v) and _is_numberish(dia_v) and _is_numberish(pul_v):
                good_rows += 1
        if good_rows >= 1:
            return r, col_map

    raise ValueError("Could not find a SmartBP table header row in the scanned area.")

def read_table_block(ws, header_row: int, col_map: Dict[str, int]) -> pd.DataFrame:
    """
    Read contiguous block under header_row until a blank Date cell (or end).
    Uses the header row values as DataFrame columns.
    """
    # Determine last column as the rightmost mapped column OR last non-empty header cell
    max_col = max(col_map.values())
    # Expand to include any adjacent headers to the right (common for Notes/Tags)
    # We'll extend until we hit a run of empty cells.
    empty_run = 0
    c = max_col
    while c <= ws.max_column and empty_run < 3:
        v = ws.cell(header_row, c).value
        if _norm_header(v):
            max_col = c
            empty_run = 0
        else:
            empty_run += 1
        c += 1

    headers = [_norm_cell_text(ws.cell(header_row, c).value) for c in range(1, max_col + 1)]

    data_rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        date_cell = ws.cell(r, col_map["date"]).value
        if date_cell is None or _norm_cell_text(date_cell) == "":
            # stop at first blank date cell
            break
        row = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        data_rows.append(row)

    df = pd.DataFrame(data_rows, columns=headers)
    return df

def convert_xlsx_to_smartbp_csv(xlsx_path: str, out_csv_path: str, sheet: Optional[str], default_source: int = 0):
    wb = load_workbook(xlsx_path, data_only=True)

    # Choose sheet
    if sheet:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")
        ws = wb[sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    header_row, col_map = find_smartbp_table(ws)
    df = read_table_block(ws, header_row, col_map)

    # Now map columns by flexible names (works with your headers too)
    def pick(*cands: str) -> Optional[str]:
        norm_map = {re.sub(r"\s+", " ", c.strip().lower()): c for c in df.columns}
        for cand in cands:
            key = re.sub(r"\s+", " ", cand.strip().lower())
            if key in norm_map:
                return norm_map[key]
        return None

    c_date = pick("Date")
    c_sys = pick("Sys (mmHg)", "Sys", "Systolic(mmHg)", "Systolic")
    c_dia = pick("Dia (mmHg)", "Dia", "Diastolic(mmHg)", "Diastolic")
    c_pulse = pick("Pulse (BPM)", "Pulse(BPM)", "Pulse", "HR", "Heart Rate")
    c_weight = pick("Weight kgs", "Weight(kgs)", "Weight (kgs)", "Weight(kgs)", "Weight")
    c_pp = pick("PP (mmHg)", "PP", "Pulse pressure(mmHg)", "PP (mmHg)")
    c_map = pick("MAP (mmHg)", "MAP", "MAP( mmHg)")
    c_notes = pick("Notes")
    c_tags = pick("Tags")

    missing = [name for name, col in [("Date", c_date), ("Sys", c_sys), ("Dia", c_dia), ("Pulse", c_pulse)] if col is None]
    if missing:
        raise ValueError(f"Missing required columns in detected table: {', '.join(missing)}")

    rows_out = []
    for _, row in df.iterrows():
        dt = _parse_datetime(row[c_date])
        date_out = _fmt_smartbp_datetime(dt)

        sys_v = _to_int_or_blank(row[c_sys])
        dia_v = _to_int_or_blank(row[c_dia])
        pulse_v = _to_int_or_blank(row[c_pulse])

        weight_v = _to_float_or_blank(row[c_weight]) if c_weight else "NA"
        if weight_v == "":
            weight_v = "NA"

        pp_tbl = _to_int_or_blank(row[c_pp]) if c_pp else ""
        map_v = _to_float_or_blank(row[c_map]) if c_map else ""

        tags_v = _split_tags(row[c_tags]) if c_tags else ""

        sys_v, dia_v, pp_out, extra_note = _check_and_fix_sys_dia(sys_v, dia_v, pp_tbl)

        user_notes = row[c_notes] if c_notes else ""
        notes_out = _build_notes(tags_v, user_notes, extra_note)

        rows_out.append([
            date_out,
            sys_v,
            dia_v,
            pulse_v,
            weight_v,
            pp_out,
            map_v,
            notes_out,
            tags_v,
            "",
            str(default_source),
        ])

    with open(out_csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(
            f,
            delimiter=",",
            quotechar='"',
            quoting=csv.QUOTE_ALL,
            lineterminator="\r\n",
            doublequote=True,
        )
        writer.writerow(SMARTBP_HEADER)
        for r in rows_out:
            r = list(r)
            r[7] = r[7].replace("\n", "\r\n")  # ensure CRLF inside Notes
            writer.writerow(r)

import os
from datetime import datetime

def main():
    ap = argparse.ArgumentParser(
        description="Convert XLSX SmartBP-like table to SmartBP-optimized CSV (auto-detect table)."
    )

    ap.add_argument("xlsx", help="Input .xlsx path")

    ap.add_argument(
        "--out",
        help="Output CSV path (optional). If omitted, a name will be generated."
    )

    ap.add_argument("--sheet", default=None, help="Sheet name (default: first sheet)")
    ap.add_argument("--source", type=int, default=0, help="SmartBP source field value (default: 0)")

    args = ap.parse_args()

    input_path = args.xlsx

    if args.out:
        out_csv = args.out
    else:
        base = os.path.splitext(os.path.basename(input_path))[0]

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        out_csv = f"{base}_smartbp_{timestamp}.csv"

    print(f"Input : {input_path}")
    print(f"Output: {out_csv}")

    convert_xlsx_to_smartbp_csv(
        input_path,
        out_csv,
        sheet=args.sheet,
        default_source=args.source
    )
if __name__ == "__main__":
    main()